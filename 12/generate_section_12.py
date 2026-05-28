"""
Section 12 – Risk Register
Generates:
  - section-12-risk-register.xlsx   (styled Excel workbook)
  - section-12-risk-register.pdf    (PDF copy with risk register + top-risk mitigation page)
"""
from __future__ import annotations

from pathlib import Path

# ── Excel imports ──────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── PDF imports ────────────────────────────────────────────────────────────────
import html as html_module
import re

from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    HRFlowable, PageBreak, Paragraph, SimpleDocTemplate,
    Spacer, Table, TableStyle,
)

ROOT = Path(__file__).resolve().parent
XLSX_OUT = ROOT / "section-12-risk-register.xlsx"
PDF_OUT  = ROOT / "section-12-risk-register.pdf"

# ──────────────────────────────────────────────────────────────────────────────
#  RISK DATA
# ──────────────────────────────────────────────────────────────────────────────
# Scales (custom, not 0-1, not dollars):
#   Impact    1–5  (1=Negligible, 2=Minor, 3=Moderate, 4=Major, 5=Critical)
#   Likelihood 1–5 (1=Rare, 2=Unlikely, 3=Possible, 4=Likely, 5=Almost Certain)
#   Risk Score = Impact × Likelihood  (range 1–25)
#   Priority band: 1-6 Low | 7-12 Medium | 13-19 High | 20-25 Critical

IMPACT_LABELS     = {1:"Negligible",2:"Minor",3:"Moderate",4:"Major",5:"Critical"}
LIKELIHOOD_LABELS = {1:"Rare",2:"Unlikely",3:"Possible",4:"Likely",5:"Almost Certain"}

def priority_band(score: int) -> str:
    if score >= 20: return "Critical"
    if score >= 13: return "High"
    if score >=  7: return "Medium"
    return "Low"

RISKS = [
    # (ID, Category, Risk Title, Description, Impact(1-5), Likelihood(1-5))
    ("R-01", "Infrastructure",
     "AWS Lab Credentials Expire Mid-Session",
     "Temporary AWS Academy credentials expire every few hours. A live deployment or demo can fail completely if credentials lapse during SAM deploy, SQS operation, or Secrets Manager access.",
     4, 5),

    ("R-02", "External Dependency",
     "OpenAI API Quota Exceeded During Exam Evaluation",
     "The evaluation Lambda calls OpenAI with no quota guard. High concurrent usage or a low-tier API key can trigger rate-limit errors, leaving reps with no coaching result after completing an exam.",
     4, 4),

    ("R-03", "Infrastructure",
     "Cognito User Pool Deleted on CloudFormation Stack Update",
     "Renaming the Cognito resource in template.yaml causes CloudFormation to delete and recreate the User Pool, permanently deleting all user accounts and invalidating all active tokens.",
     5, 3),

    ("R-04", "Performance",
     "DynamoDB Full-Table Scan Degrades Under Load",
     "The dashboard, persona list, scenario list, and exam scenario list all use Scan. Under heavy load or a large dataset the scan latency can exceed Lambda timeouts and degrade the manager experience.",
     3, 3),

    ("R-05", "Reliability",
     "SQS Delayed Message Not Delivered During Exam",
     "If the SQS delivery for an exam issue is late or dropped, the issue will not become visible on time. The pulse endpoint mitigates this but only when polled; issues may still appear late.",
     3, 2),

    ("R-06", "External Dependency",
     "OpenAI Model Deprecated or API Schema Change",
     "The backend hardcodes the OpenAI Responses API endpoint and JSON schema. A model deprecation or API version change would silently break issue generation and evaluation without code changes.",
     4, 3),

    ("R-07", "Security",
     "CORS Wildcard Origin Allows Unauthorized Frontend",
     "The current CORS configuration allows any origin. Any web page can make credentialed calls to the API if it possesses a valid Cognito token, widening the attack surface for token-stealing exploits.",
     3, 3),

    ("R-08", "Data Integrity",
     "Demo Issues Delivered to Reps Without Manager Awareness",
     "When OpenAI fails during issue generation the backend silently stores demo issues. If the manager does not check the generationWarning field, reps take exams on scripted dummy content without knowing.",
     3, 4),

    ("R-09", "Data Integrity",
     "DynamoDB Item Size Limit Exceeded by Large Scenario",
     "All generated issues are stored inside one DynamoDB item on the Scenarios table. At the maximum 20 issues with long messages the item may approach or exceed DynamoDB's 400 KB item size limit.",
     2, 2),

    ("R-10", "Reliability",
     "Exam Clock Skew Causes Premature or Late Issue Release",
     "Issue release timing is calculated server-side using Lambda clock. Large clock skew between Lambda invocations or between SQS and Lambda can cause issues to be revealed out of order or late.",
     3, 2),

    ("R-11", "Operations",
     "Manager Archives All Scenarios Before a Scheduled Exam",
     "The archive action has no confirmation guard in the backend. A manager can accidentally archive all published scenarios, making the exam list empty for all reps until scenarios are republished.",
     4, 2),

    ("R-12", "Security",
     "Refresh Token Exposure Enables Long-Lived Session Hijack",
     "Refresh tokens are valid for 30 days and are stored in browser local storage without HttpOnly protection. XSS on the frontend or local storage inspection can expose the token to an attacker.",
     4, 3),
]

# Augment with computed fields
RISK_ROWS = []
for row in RISKS:
    rid, cat, title, desc, impact, likelihood = row
    score = impact * likelihood
    band  = priority_band(score)
    RISK_ROWS.append({
        "id": rid, "category": cat, "title": title, "description": desc,
        "impact": impact, "impact_label": IMPACT_LABELS[impact],
        "likelihood": likelihood, "likelihood_label": LIKELIHOOD_LABELS[likelihood],
        "score": score, "band": band,
    })

RISK_ROWS.sort(key=lambda r: r["score"], reverse=True)
TOP_RISK = RISK_ROWS[0]

# ──────────────────────────────────────────────────────────────────────────────
#  EXCEL GENERATION
# ──────────────────────────────────────────────────────────────────────────────
BAND_COLORS = {
    "Critical": ("C0392B", "FFFFFF"),
    "High":     ("E67E22", "FFFFFF"),
    "Medium":   ("F1C40F", "1a1a1a"),
    "Low":      ("27AE60", "FFFFFF"),
}

def side(style="thin", color="C5DDD7"):
    return Side(style=style, color=color)

THIN_BORDER = Border(left=side(), right=side(), top=side(), bottom=side())

def header_fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def build_excel():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Risk Register ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Risk Register"

    # Title block
    ws.merge_cells("A1:K1")
    ws["A1"] = "SalesOps AI — Risk Register"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = header_fill("1A5C4A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:K2")
    ws["A2"] = (
        "Impact scale: 1=Negligible · 2=Minor · 3=Moderate · 4=Major · 5=Critical    "
        "Likelihood scale: 1=Rare · 2=Unlikely · 3=Possible · 4=Likely · 5=Almost Certain    "
        "Risk Score = Impact × Likelihood  (max 25)"
    )
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="17324D")
    ws["A2"].fill = header_fill("E8F5F1")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Column headers
    HEADERS = [
        "Risk ID", "Category", "Risk Title", "Description",
        "Impact\n(1–5)", "Impact Label",
        "Likelihood\n(1–5)", "Likelihood Label",
        "Risk Score\n(1–25)", "Priority Band",
        "Owner",
    ]
    COL_WIDTHS = [8, 18, 32, 58, 9, 16, 12, 18, 11, 14, 14]

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = header_fill("17324D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 30

    # Data rows
    owners = {
        "Infrastructure": "DevOps Lead",
        "External Dependency": "Backend Dev",
        "Performance": "Backend Dev",
        "Reliability": "Backend Dev",
        "Security": "Security Lead",
        "Data Integrity": "Backend Dev",
        "Operations": "Project Manager",
    }

    for row_idx, r in enumerate(RISK_ROWS, start=4):
        bg_hex, fg_hex = BAND_COLORS[r["band"]]
        band_fill  = PatternFill("solid", fgColor=bg_hex)
        band_font  = Font(name="Calibri", bold=True, size=10, color=fg_hex)
        score_fill = PatternFill("solid", fgColor=bg_hex)
        score_font = Font(name="Calibri", bold=True, size=11, color=fg_hex)

        row_bg = "F5FAF8" if row_idx % 2 == 0 else "FFFFFF"
        normal_fill = PatternFill("solid", fgColor=row_bg)
        normal_font = Font(name="Calibri", size=10)

        values = [
            r["id"], r["category"], r["title"], r["description"],
            r["impact"], r["impact_label"],
            r["likelihood"], r["likelihood_label"],
            r["score"], r["band"],
            owners.get(r["category"], "Team Lead"),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col_idx == 9:  # Score
                cell.fill = score_fill
                cell.font = score_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 10:  # Band
                cell.fill = band_fill
                cell.font = band_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (5, 7):  # Numeric scales
                cell.fill = normal_fill
                cell.font = Font(name="Calibri", bold=True, size=10, color="17324D")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 1:  # ID
                cell.fill = normal_fill
                cell.font = Font(name="Calibri", bold=True, size=10, color="1A5C4A")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = normal_fill
                cell.font = normal_font

        ws.row_dimensions[row_idx].height = 52

    # Freeze panes
    ws.freeze_panes = "A4"

    # Auto-filter
    ws.auto_filter.ref = f"A3:K{3 + len(RISK_ROWS)}"

    # ── Sheet 2: Priority Matrix (heatmap summary) ─────────────────────────────
    ws2 = wb.create_sheet("Priority Matrix")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "Risk Priority Matrix — Likelihood vs Impact"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill = header_fill("1A5C4A")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    # Axes
    matrix_row_labels = ["5 Almost\nCertain", "4 Likely", "3 Possible", "2 Unlikely", "1 Rare"]
    matrix_col_labels = ["", "1\nNegligible", "2\nMinor", "3\nModerate", "4\nMajor", "5\nCritical"]

    for col_idx, label in enumerate(matrix_col_labels, start=1):
        cell = ws2.cell(row=2, column=col_idx, value=label)
        if col_idx > 1:
            cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            cell.fill = header_fill("17324D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = 14
    ws2.row_dimensions[2].height = 30

    # Matrix cells
    for r_idx, row_label in enumerate(matrix_row_labels, start=3):
        likelihood = 5 - (r_idx - 3)
        label_cell = ws2.cell(row=r_idx, column=1, value=row_label)
        label_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        label_cell.fill = header_fill("17324D")
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        label_cell.border = THIN_BORDER
        ws2.row_dimensions[r_idx].height = 48

        for c_idx in range(1, 6):
            impact = c_idx
            score  = impact * likelihood
            band   = priority_band(score)
            bg, fg = BAND_COLORS[band]

            # Find risks at this cell
            matching = [r["id"] for r in RISK_ROWS if r["impact"] == impact and r["likelihood"] == likelihood]
            cell_val = ", ".join(matching) if matching else str(score)

            cell = ws2.cell(row=r_idx, column=c_idx + 1, value=cell_val)
            cell.fill  = PatternFill("solid", fgColor=bg)
            cell.font  = Font(name="Calibri", bold=bool(matching), size=10, color=fg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    # Scale legend
    legend_row = 9
    ws2.merge_cells(f"A{legend_row}:H{legend_row}")
    ws2[f"A{legend_row}"] = "Legend:"
    ws2[f"A{legend_row}"].font = Font(name="Calibri", bold=True, size=11)
    legend_row += 1
    for band, (bg, fg) in BAND_COLORS.items():
        ws2.merge_cells(f"A{legend_row}:B{legend_row}")
        cell = ws2[f"A{legend_row}"]
        cell.value = band
        cell.fill  = PatternFill("solid", fgColor=bg)
        cell.font  = Font(name="Calibri", bold=True, size=11, color=fg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        range_text = {"Critical": "Score 20–25", "High": "Score 13–19",
                      "Medium": "Score 7–12", "Low": "Score 1–6"}[band]
        ws2.merge_cells(f"C{legend_row}:D{legend_row}")
        ws2[f"C{legend_row}"] = range_text
        ws2[f"C{legend_row}"].font = Font(name="Calibri", size=10)
        ws2.row_dimensions[legend_row].height = 22
        legend_row += 1

    # ── Sheet 3: Top Risk Mitigation ───────────────────────────────────────────
    ws3 = wb.create_sheet("Top Risk Mitigation")
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 70

    ws3.merge_cells("A1:B1")
    ws3["A1"] = f"Detailed Mitigation Plan — {TOP_RISK['id']}: {TOP_RISK['title']}"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = header_fill("C0392B")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws3.row_dimensions[1].height = 36

    detail_rows = [
        ("Risk ID",           TOP_RISK["id"]),
        ("Category",          TOP_RISK["category"]),
        ("Risk Title",        TOP_RISK["title"]),
        ("Description",       TOP_RISK["description"]),
        ("Impact",            f"{TOP_RISK['impact']} – {TOP_RISK['impact_label']}"),
        ("Likelihood",        f"{TOP_RISK['likelihood']} – {TOP_RISK['likelihood_label']}"),
        ("Risk Score",        f"{TOP_RISK['score']} / 25  ({TOP_RISK['band']})"),
        ("Owner",             "DevOps Lead / Project Manager"),
        ("", ""),
        ("Root Cause",
         "AWS Academy provides temporary STS credentials that expire every 4-8 hours. "
         "The project requires these credentials for every SAM deploy, Lambda invocation test, "
         "DynamoDB CLI operation, and Secrets Manager rotation. There is no automated refresh "
         "mechanism in the repository — the developer must manually copy new credentials before each session."),
        ("Impact Detail",
         "If credentials expire during a SAM deployment, CloudFormation may enter a partial update state "
         "requiring manual rollback. If they expire during an exam demo, SQS and DynamoDB operations will "
         "fail with AuthorizationError, ending the exam session abruptly. Any Lambda invocation "
         "requiring Secrets Manager (issue generation, evaluation) will also fail, leaving reps "
         "without scoring results."),
        ("", ""),
        ("Mitigation — Immediate",
         "1. Add a pre-session checklist to the admin guide requiring credential refresh before every lab session.\n"
         "2. Add a startup script (scripts/check-creds.sh) that runs `aws sts get-caller-identity` and exits "
         "with an error message if credentials are expired or missing.\n"
         "3. Display a credential expiry warning in the README and admin guide with the exact refresh procedure."),
        ("Mitigation — Short Term",
         "1. Wrap the SAM deploy npm script to call `aws sts get-caller-identity` first and abort if it fails.\n"
         "2. Add a GitHub Actions workflow step that validates credentials before any deploy step.\n"
         "3. Store credential refresh instructions as a pinned note in the team project board."),
        ("Mitigation — Long Term",
         "1. Migrate from AWS Academy to a dedicated AWS account where long-lived IAM roles with "
         "instance profiles eliminate per-session credential rotation entirely.\n"
         "2. Use AWS CLI named profiles with SSO to allow automated token refresh without manual copy-paste.\n"
         "3. Implement infrastructure-level monitoring (CloudWatch alarm on Lambda auth errors) "
         "to alert the team immediately when credentials expire unexpectedly during a session."),
        ("Residual Risk After Mitigation",
         "Impact reduced to 2 (Minor) — automated pre-session check catches the issue before it affects "
         "end users. Likelihood reduced to 3 (Possible) — manual step still required in Academy environment. "
         "Residual Score: 2 × 3 = 6 (Low)."),
        ("Acceptance Criteria",
         "Risk is considered mitigated when: (a) credential check script exists and passes in CI, "
         "(b) admin guide section 08 includes the refresh checklist, (c) no deployment has failed "
         "due to expired credentials in the last 3 lab sessions."),
    ]

    for row_idx, (label, value) in enumerate(detail_rows, start=2):
        if not label and not value:
            ws3.row_dimensions[row_idx].height = 10
            continue
        label_cell = ws3.cell(row=row_idx, column=1, value=label)
        value_cell = ws3.cell(row=row_idx, column=2, value=value)

        if label:
            label_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            label_cell.fill = header_fill("17324D")
            label_cell.border = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="top",
                                             wrap_text=True, indent=1)
        value_cell.font = Font(name="Calibri", size=10)
        value_cell.fill = PatternFill("solid", fgColor="F5FAF8")
        value_cell.border = THIN_BORDER
        value_cell.alignment = Alignment(horizontal="left", vertical="top",
                                         wrap_text=True, indent=1)

        if label == "Risk Score":
            value_cell.fill = PatternFill("solid", fgColor="C0392B")
            value_cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        elif label in ("Mitigation — Immediate", "Mitigation — Short Term", "Mitigation — Long Term"):
            value_cell.fill = PatternFill("solid", fgColor="EAF5F1")

        ws3.row_dimensions[row_idx].height = max(
            22, min(120, 14 * (value.count("\n") + len(value) // 70 + 1))
        )

    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT}")

# ──────────────────────────────────────────────────────────────────────────────
#  PDF GENERATION
# ──────────────────────────────────────────────────────────────────────────────
TEAL      = colors.HexColor("#1a5c4a")
TEAL_L    = colors.HexColor("#c8e6df")
SLATE     = colors.HexColor("#17324d")
MUTED_C   = colors.HexColor("#5b6570")
WHITE     = colors.white
ROW_A     = colors.HexColor("#f5faf8")
ROW_B     = colors.white

PDF_BAND_COLORS = {
    "Critical": (colors.HexColor("#c0392b"), WHITE),
    "High":     (colors.HexColor("#e67e22"), WHITE),
    "Medium":   (colors.HexColor("#f1c40f"), colors.HexColor("#1a1a1a")),
    "Low":      (colors.HexColor("#27ae60"), WHITE),
}

def pdf_styles():
    base = getSampleStyleSheet()
    return {
        "title":  ParagraphStyle("T",  parent=base["Title"],
                                 fontName="Helvetica-Bold", fontSize=20, leading=26,
                                 textColor=SLATE, spaceAfter=8),
        "h2":     ParagraphStyle("H2", parent=base["Heading2"],
                                 fontName="Helvetica-Bold", fontSize=13, leading=17,
                                 textColor=TEAL, spaceBefore=16, spaceAfter=5),
        "h3":     ParagraphStyle("H3", parent=base["Heading3"],
                                 fontName="Helvetica-Bold", fontSize=11, leading=14,
                                 textColor=SLATE, spaceBefore=10, spaceAfter=4),
        "body":   ParagraphStyle("B",  parent=base["BodyText"],
                                 fontName="Helvetica", fontSize=9.0, leading=11.8,
                                 textColor=colors.HexColor("#20252b"), spaceAfter=4),
        "small":  ParagraphStyle("Sm", parent=base["BodyText"],
                                 fontName="Helvetica", fontSize=8.0, leading=10,
                                 textColor=MUTED_C, spaceAfter=3),
        "bullet": ParagraphStyle("Bl", parent=base["BodyText"],
                                 fontName="Helvetica", fontSize=9.0, leading=11.5,
                                 leftIndent=14, firstLineIndent=-7,
                                 textColor=colors.HexColor("#20252b"), spaceAfter=2.5),
        "th":     ParagraphStyle("TH", parent=base["BodyText"],
                                 fontName="Helvetica-Bold", fontSize=8.0, leading=10, textColor=WHITE),
        "td":     ParagraphStyle("TD", parent=base["BodyText"],
                                 fontName="Helvetica", fontSize=8.0, leading=10.2,
                                 textColor=colors.HexColor("#20252b")),
        "td_b":   ParagraphStyle("TDB",parent=base["BodyText"],
                                 fontName="Helvetica-Bold", fontSize=8.0, leading=10.2,
                                 textColor=colors.HexColor("#17324d")),
        "label":  ParagraphStyle("LB", parent=base["BodyText"],
                                 fontName="Helvetica-Bold", fontSize=9.0, leading=11,
                                 textColor=WHITE),
    }

def hdr_ftr(canvas, doc):
    canvas.saveState()
    w, h = LETTER
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(MUTED_C)
    canvas.drawString(doc.leftMargin, h - 0.40*inch, "Group - E  |  SalesOps AI  |  Section 12")
    canvas.drawRightString(w - doc.rightMargin, 0.40*inch, f"Page {doc.page}")
    canvas.setStrokeColor(TEAL_L)
    canvas.setLineWidth(0.5)
    canvas.line(doc.leftMargin, h - 0.48*inch, w - doc.rightMargin, h - 0.48*inch)
    canvas.restoreState()

def build_pdf():
    S = pdf_styles()
    story = []

    # ── Cover / intro ──────────────────────────────────────────────────────────
    story.append(Paragraph("Section 12 – Risk Register", S["title"]))
    story.append(Paragraph("Prepared: May 27, 2026", S["small"]))
    story.append(Spacer(1, 0.08*inch))
    story.append(Paragraph(
        "This register identifies 12 risks across infrastructure, external dependencies, "
        "performance, reliability, security, data integrity, and operations. "
        "Risks are scored using custom integer scales (not 0–1, not dollar values).",
        S["body"]))
    story.append(Spacer(1, 0.06*inch))

    # Scale legend table
    legend_data = [
        [Paragraph("Scale", S["th"]),
         Paragraph("Values", S["th"]),
         Paragraph("Meaning", S["th"])],
        [Paragraph("Impact (1–5)", S["td_b"]),
         Paragraph("1 · 2 · 3 · 4 · 5", S["td"]),
         Paragraph("Negligible · Minor · Moderate · Major · Critical", S["td"])],
        [Paragraph("Likelihood (1–5)", S["td_b"]),
         Paragraph("1 · 2 · 3 · 4 · 5", S["td"]),
         Paragraph("Rare · Unlikely · Possible · Likely · Almost Certain", S["td"])],
        [Paragraph("Risk Score (1–25)", S["td_b"]),
         Paragraph("Impact × Likelihood", S["td"]),
         Paragraph("1–6 Low  |  7–12 Medium  |  13–19 High  |  20–25 Critical", S["td"])],
    ]
    avail = 6.4*inch
    lt = Table(legend_data, colWidths=[1.4*inch, 1.6*inch, avail-3.0*inch])
    lt.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), SLATE),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[ROW_A, ROW_B, ROW_A]),
        ("GRID",(0,0),(-1,-1),0.35,TEAL_L),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(lt)
    story.append(Spacer(1, 0.12*inch))

    # ── Risk Register table ────────────────────────────────────────────────────
    story.append(Paragraph("Risk Register", S["h2"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=TEAL_L, spaceAfter=6))

    avail = 6.4*inch
    col_widths = [0.42*inch, 0.82*inch, 1.55*inch, 2.15*inch,
                  0.50*inch, 0.50*inch, 0.76*inch]

    headers = ["ID","Category","Risk Title","Description","Impact","Like-\nlihood","Score\n& Band"]
    header_row = [Paragraph(h, S["th"]) for h in headers]

    table_data = [header_row]
    table_styles = [
        ("BACKGROUND", (0,0),(-1,0), SLATE),
        ("GRID",(0,0),(-1,-1),0.35,TEAL_L),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
    ]

    for r_idx, r in enumerate(RISK_ROWS, start=1):
        bg_c, fg_c = PDF_BAND_COLORS[r["band"]]
        score_text = f'{r["score"]}\n{r["band"]}'
        score_style = ParagraphStyle("SC", fontName="Helvetica-Bold", fontSize=8.0,
                                     leading=10, textColor=fg_c, alignment=1)
        row = [
            Paragraph(r["id"],        S["td_b"]),
            Paragraph(r["category"],  S["td"]),
            Paragraph(r["title"],     S["td_b"]),
            Paragraph(r["description"][:220] + ("…" if len(r["description"])>220 else ""), S["td"]),
            Paragraph(f'{r["impact"]}\n{r["impact_label"]}',       S["td"]),
            Paragraph(f'{r["likelihood"]}\n{r["likelihood_label"]}', S["td"]),
            Paragraph(score_text, score_style),
        ]
        table_data.append(row)
        row_bg = ROW_A if r_idx % 2 else ROW_B
        table_styles.append(("BACKGROUND", (0, r_idx),  (5, r_idx), row_bg))
        table_styles.append(("BACKGROUND", (6, r_idx),  (6, r_idx), bg_c))

    rt = Table(table_data, colWidths=col_widths, repeatRows=1)
    rt.setStyle(TableStyle(table_styles))
    story.append(rt)

    # ── Page break → Top Risk Mitigation ──────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph(
        f"Top Risk Mitigation Plan — {TOP_RISK['id']}: {TOP_RISK['title']}", S["h2"]))
    story.append(HRFlowable(width="100%", thickness=1.0, color=colors.HexColor("#c0392b"),
                             spaceAfter=8))

    # Summary card
    bg_c, fg_c = PDF_BAND_COLORS[TOP_RISK["band"]]
    card_data = [
        [Paragraph("Risk ID",    S["label"]), Paragraph(TOP_RISK["id"],       S["td"])],
        [Paragraph("Category",   S["label"]), Paragraph(TOP_RISK["category"], S["td"])],
        [Paragraph("Impact",     S["label"]),
         Paragraph(f'{TOP_RISK["impact"]} – {TOP_RISK["impact_label"]}', S["td"])],
        [Paragraph("Likelihood", S["label"]),
         Paragraph(f'{TOP_RISK["likelihood"]} – {TOP_RISK["likelihood_label"]}', S["td"])],
        [Paragraph("Risk Score", S["label"]),
         Paragraph(f'{TOP_RISK["score"]} / 25  →  {TOP_RISK["band"]}',
                   ParagraphStyle("RS", fontName="Helvetica-Bold", fontSize=10,
                                  textColor=fg_c))],
        [Paragraph("Owner",      S["label"]), Paragraph("DevOps Lead / Project Manager", S["td"])],
    ]
    ct = Table(card_data, colWidths=[1.5*inch, 4.9*inch])
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(0,-1), SLATE),
        ("BACKGROUND", (1,0),(1,3),  ROW_A),
        ("BACKGROUND", (1,4),(1,4),  colors.HexColor("#fde8e6")),
        ("BACKGROUND", (1,5),(1,5),  ROW_A),
        ("GRID",(0,0),(-1,-1),0.4,TEAL_L),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(ct)
    story.append(Spacer(1, 0.12*inch))

    # Full description
    story.append(Paragraph("Description", S["h3"]))
    story.append(Paragraph(TOP_RISK["description"], S["body"]))
    story.append(Spacer(1, 0.06*inch))

    story.append(Paragraph("Root Cause", S["h3"]))
    story.append(Paragraph(
        "AWS Academy provides temporary STS credentials that expire every 4–8 hours. "
        "The project requires these credentials for every SAM deploy, Lambda invocation test, "
        "DynamoDB CLI operation, and Secrets Manager rotation. There is no automated refresh "
        "mechanism in the repository — the developer must manually copy new credentials before each session.",
        S["body"]))

    story.append(Paragraph("Impact Detail", S["h3"]))
    story.append(Paragraph(
        "If credentials expire during a SAM deployment, CloudFormation may enter a partial update state "
        "requiring manual rollback. If they expire during an exam demo, SQS and DynamoDB operations "
        "will fail with AuthorizationError, ending the exam session abruptly. Any Lambda invocation "
        "requiring Secrets Manager (issue generation, evaluation) will also fail, leaving reps without "
        "scoring results.",
        S["body"]))
    story.append(Spacer(1, 0.06*inch))

    # Mitigation sections
    mitigations = [
        ("Mitigation — Immediate Actions",
         colors.HexColor("#fde8e6"), colors.HexColor("#c0392b"),
         [
             "Add a pre-session checklist to the admin guide requiring credential refresh before every lab session.",
             "Add a startup script (scripts/check-creds.sh) that runs aws sts get-caller-identity and exits with an error message if credentials are expired or missing.",
             "Display a credential expiry warning in the README with the exact refresh procedure.",
         ]),
        ("Mitigation — Short-Term Actions",
         colors.HexColor("#fff4e6"), colors.HexColor("#c87a10"),
         [
             "Wrap the SAM deploy npm script to call aws sts get-caller-identity first and abort if it fails.",
             "Add a GitHub Actions workflow step that validates credentials before any deploy step.",
             "Store credential refresh instructions as a pinned note in the team project board.",
         ]),
        ("Mitigation — Long-Term Actions",
         colors.HexColor("#f0f9f6"), TEAL,
         [
             "Migrate from AWS Academy to a dedicated AWS account where long-lived IAM roles with instance profiles eliminate per-session credential rotation entirely.",
             "Use AWS CLI named profiles with SSO to allow automated token refresh without manual copy-paste.",
             "Implement CloudWatch alarm on Lambda AuthorizationError to alert the team immediately when credentials expire during a session.",
         ]),
    ]

    for mit_title, bg, fg, points in mitigations:
        title_para = Paragraph(mit_title, ParagraphStyle(
            "MT", fontName="Helvetica-Bold", fontSize=10, leading=13,
            textColor=fg, spaceBefore=10, spaceAfter=4))
        story.append(title_para)
        for pt in points:
            story.append(Paragraph(pt, S["bullet"], bulletText="•"))
        story.append(Spacer(1, 0.04*inch))

    story.append(Spacer(1, 0.06*inch))

    # Residual risk table
    story.append(Paragraph("Residual Risk After Full Mitigation", S["h3"]))
    res_data = [
        [Paragraph(h, S["th"]) for h in ["Dimension","Before","After","Change"]],
        [Paragraph("Impact",    S["td_b"]),
         Paragraph(f'{TOP_RISK["impact"]} – {TOP_RISK["impact_label"]}', S["td"]),
         Paragraph("2 – Minor",     S["td"]),
         Paragraph("↓ −2", ParagraphStyle("CH",fontName="Helvetica-Bold",fontSize=9,textColor=TEAL))],
        [Paragraph("Likelihood", S["td_b"]),
         Paragraph(f'{TOP_RISK["likelihood"]} – {TOP_RISK["likelihood_label"]}', S["td"]),
         Paragraph("3 – Possible",  S["td"]),
         Paragraph("↓ −2", ParagraphStyle("CH",fontName="Helvetica-Bold",fontSize=9,textColor=TEAL))],
        [Paragraph("Risk Score", S["td_b"]),
         Paragraph(f'{TOP_RISK["score"]} — Critical', S["td"]),
         Paragraph("6 — Low",       S["td"]),
         Paragraph("↓ −14", ParagraphStyle("CH",fontName="Helvetica-Bold",fontSize=9,
                                            textColor=TEAL))],
    ]
    avail = 6.4*inch
    res_t = Table(res_data, colWidths=[1.6*inch, 1.9*inch, 1.5*inch, 1.4*inch])
    res_t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),SLATE),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[ROW_A, ROW_B, colors.HexColor("#eaf5f1")]),
        ("GRID",(0,0),(-1,-1),0.35,TEAL_L),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(res_t)
    story.append(Spacer(1, 0.08*inch))

    story.append(Paragraph("Acceptance Criteria", S["h3"]))
    for criterion in [
        "Credential check script exists and passes in CI before every deployment.",
        "Admin guide section 08 includes the credential refresh checklist.",
        "No deployment has failed due to expired credentials in the last 3 consecutive lab sessions.",
    ]:
        story.append(Paragraph(criterion, S["bullet"], bulletText="✓"))

    doc = SimpleDocTemplate(
        str(PDF_OUT), pagesize=LETTER,
        rightMargin=0.70*inch, leftMargin=0.70*inch,
        topMargin=0.72*inch, bottomMargin=0.64*inch,
        title="Group - E | SalesOps AI | Section 12",
        author="Group E",
        subject="Section 12 Risk Register",
    )
    doc.build(story, onFirstPage=hdr_ftr, onLaterPages=hdr_ftr)
    print(f"Wrote {PDF_OUT}")


if __name__ == "__main__":
    build_excel()
    build_pdf()
